#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import os
import sqlite3
from typing import Dict, List, Optional, Sequence, Tuple

import trueskill as ts
import math


DB_DEFAULT_PATH = os.path.join(os.path.dirname(__file__), "data", "ratings.db")


def ensure_parent_dir(path: str) -> None:
	parent = os.path.dirname(path)
	if parent and not os.path.exists(parent):
		os.makedirs(parent, exist_ok=True)


def open_db(db_path: str) -> sqlite3.Connection:
	ensure_parent_dir(db_path)
	conn = sqlite3.connect(db_path)
	conn.row_factory = sqlite3.Row
	return conn


def table_has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
	cur = conn.cursor()
	cur.execute(f"PRAGMA table_info({table})")
	for row in cur.fetchall():
		if str(row[1]).lower() == column.lower():
			return True
	return False


def init_db(conn: sqlite3.Connection) -> None:
	cur = conn.cursor()
	# players with sex, weight
	cur.execute(
		"""
		CREATE TABLE IF NOT EXISTS players (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			name TEXT NOT NULL UNIQUE COLLATE NOCASE,
			sex TEXT,              -- 'M' | 'F' | NULL
			weight REAL            -- pounds; NULL if unknown
		);
		"""
	)
	# ratings keyed by player+hand only
	cur.execute(
		"""
		CREATE TABLE IF NOT EXISTS ratings (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			player_id INTEGER NOT NULL,
			context_hand TEXT NOT NULL,           -- 'RH' | 'LH'
			mu REAL NOT NULL,
			sigma REAL NOT NULL,
			updated_at TEXT NOT NULL,
			UNIQUE(player_id, context_hand),
			FOREIGN KEY(player_id) REFERENCES players(id)
		);
		"""
	)
	# events (keep minimal context: hand only)
	cur.execute(
		"""
		CREATE TABLE IF NOT EXISTS events (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			event_type TEXT NOT NULL,            -- 'tournament' | 'supermatch'
			name TEXT NOT NULL,
			date TEXT NOT NULL,
			context_hand TEXT NOT NULL          -- 'RH' | 'LH'
		);
		"""
	)
	cur.execute(
		"""
		CREATE TABLE IF NOT EXISTS supermatches (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			event_id INTEGER NOT NULL,
			a_player_id INTEGER NOT NULL,
			b_player_id INTEGER NOT NULL,
			a_wins INTEGER NOT NULL,
			b_wins INTEGER NOT NULL,
			best_of INTEGER NOT NULL,
			FOREIGN KEY(event_id) REFERENCES events(id)
		);
		"""
	)
	cur.execute(
		"""
		CREATE TABLE IF NOT EXISTS tournament_results (
			id INTEGER PRIMARY KEY AUTOINCREMENT,
			event_id INTEGER NOT NULL,
			player_id INTEGER NOT NULL,
			placement INTEGER NOT NULL,
			FOREIGN KEY(event_id) REFERENCES events(id)
		);
		"""
	)
	conn.commit()
	migrate_schema(conn)


def migrate_schema(conn: sqlite3.Connection) -> None:
	cur = conn.cursor()
	# Ensure sex and weight exist on players
	if not table_has_column(conn, 'players', 'sex'):
		cur.execute("ALTER TABLE players ADD COLUMN sex TEXT")
	if not table_has_column(conn, 'players', 'weight'):
		cur.execute("ALTER TABLE players ADD COLUMN weight REAL")
	conn.commit()
	# If old ratings had division/class, collapse to new ratings
	try:
		cur.execute("PRAGMA table_info(ratings)")
		cols = [str(r[1]).lower() for r in cur.fetchall()]
		if 'context_division'.lower() in cols or 'context_class'.lower() in cols:
			cur.execute(
				"""
				CREATE TABLE IF NOT EXISTS ratings_new (
					id INTEGER PRIMARY KEY AUTOINCREMENT,
					player_id INTEGER NOT NULL,
					context_hand TEXT NOT NULL,
					mu REAL NOT NULL,
					sigma REAL NOT NULL,
					updated_at TEXT NOT NULL,
					UNIQUE(player_id, context_hand)
				)
				"""
			)
			# pick the most recent row per player+hand
			cur.execute(
				"""
				INSERT OR REPLACE INTO ratings_new(player_id, context_hand, mu, sigma, updated_at)
				SELECT player_id, context_hand,
				       mu, sigma, MAX(updated_at) as updated_at
				FROM ratings
				GROUP BY player_id, context_hand
				"""
			)
			cur.execute("DROP TABLE ratings")
			cur.execute("ALTER TABLE ratings_new RENAME TO ratings")
			conn.commit()
	except Exception:
		pass
	# If old events had extra required columns, rebuild to minimal schema
	try:
		cur.execute("PRAGMA table_info(events)")
		ev_cols = [str(r[1]).lower() for r in cur.fetchall()]
		if 'context_division' in ev_cols or 'context_class' in ev_cols:
			cur.execute(
				"""
				CREATE TABLE IF NOT EXISTS events_new (
					id INTEGER PRIMARY KEY AUTOINCREMENT,
					event_type TEXT NOT NULL,
					name TEXT NOT NULL,
					date TEXT NOT NULL,
					context_hand TEXT NOT NULL
				)
				"""
			)
			cur.execute(
				"""
				INSERT INTO events_new(id, event_type, name, date, context_hand)
				SELECT id, event_type, name, date, context_hand FROM events
				"""
			)
			cur.execute("DROP TABLE events")
			cur.execute("ALTER TABLE events_new RENAME TO events")
			conn.commit()
	except Exception:
		pass


class RatingsRepository:
	def __init__(self, conn: sqlite3.Connection) -> None:
		self.conn = conn

	def get_or_create_player(self, name: str) -> int:
		name = name.strip()
		cur = self.conn.cursor()
		cur.execute("SELECT id FROM players WHERE name = ?", (name,))
		row = cur.fetchone()
		if row:
			return int(row[0])
		cur.execute("INSERT INTO players(name) VALUES(?)", (name,))
		self.conn.commit()
		return int(cur.lastrowid)

	def find_player(self, name: str) -> Optional[int]:
		cur = self.conn.cursor()
		cur.execute("SELECT id FROM players WHERE name = ? COLLATE NOCASE", (name.strip(),))
		row = cur.fetchone()
		return int(row[0]) if row else None

	def set_player_attrs(self, player_id: int, sex: Optional[str] = None, weight: Optional[float] = None) -> None:
		cur = self.conn.cursor()
		sets = []
		params: List[object] = []
		if sex is not None:
			sets.append("sex = ?")
			params.append(sex)
		if weight is not None:
			sets.append("weight = ?")
			params.append(weight)
		if sets:
			params.append(player_id)
			cur.execute(f"UPDATE players SET {', '.join(sets)} WHERE id = ?", params)
			self.conn.commit()

	def get_rating(self, player_id: int, hand: str, env: ts.TrueSkill) -> ts.Rating:
		cur = self.conn.cursor()
		cur.execute(
			"""
			SELECT mu, sigma FROM ratings
			WHERE player_id = ? AND context_hand = ?
			""",
			(player_id, hand),
		)
		row = cur.fetchone()
		if row:
			return ts.Rating(mu=float(row[0]), sigma=float(row[1]))
		# default
		return env.Rating()

	def update_rating(self, player_id: int, hand: str, rating: ts.Rating) -> None:
		cur = self.conn.cursor()
		now = dt.datetime.utcnow().isoformat()
		cur.execute(
			"""
			INSERT INTO ratings(player_id, context_hand, mu, sigma, updated_at)
			VALUES(?,?,?,?,?)
			ON CONFLICT(player_id, context_hand)
			DO UPDATE SET mu=excluded.mu, sigma=excluded.sigma, updated_at=excluded.updated_at
			""",
			(player_id, hand, float(rating.mu), float(rating.sigma), now),
		)
		self.conn.commit()

	def create_event(self, event_type: str, name: str, date_iso: str, hand: str) -> int:
		cur = self.conn.cursor()
		cur.execute(
			"INSERT INTO events(event_type, name, date, context_hand) VALUES(?,?,?,?)",
			(event_type, name, date_iso, hand),
		)
		self.conn.commit()
		return int(cur.lastrowid)

	def add_supermatch_row(self, event_id: int, a_player_id: int, b_player_id: int, a_wins: int, b_wins: int, best_of: int) -> None:
		cur = self.conn.cursor()
		cur.execute(
			"INSERT INTO supermatches(event_id, a_player_id, b_player_id, a_wins, b_wins, best_of) VALUES(?,?,?,?,?,?)",
			(event_id, a_player_id, b_player_id, a_wins, b_wins, best_of),
		)
		self.conn.commit()

	def add_tournament_result(self, event_id: int, player_id: int, placement: int) -> None:
		cur = self.conn.cursor()
		cur.execute(
			"INSERT INTO tournament_results(event_id, player_id, placement) VALUES(?,?,?)",
			(event_id, player_id, placement),
		)
		self.conn.commit()

	def fetch_ratings(self, hand: Optional[str]) -> List[sqlite3.Row]:
		cur = self.conn.cursor()
		q = [
			"SELECT p.name, r.context_hand AS hand, r.mu, r.sigma, (r.mu - 3*r.sigma) AS conservative, r.updated_at",
			"FROM ratings r JOIN players p ON p.id = r.player_id",
		]
		conds = []
		params: List[object] = []
		if hand:
			conds.append("r.context_hand = ?")
			params.append(hand)
		if conds:
			q.append("WHERE " + " AND ".join(conds))
		q.append("ORDER BY conservative DESC")
		cur.execute("\n".join(q), params)
		return list(cur.fetchall())

	def fetch_events(self, limit: int = 20) -> List[sqlite3.Row]:
		cur = self.conn.cursor()
		cur.execute("SELECT id, event_type, name, date, context_hand FROM events ORDER BY id DESC LIMIT ?", (limit,))
		return list(cur.fetchall())

	def fetch_player_overview(self, player_name: str) -> Dict[str, object]:
		pid = self.find_player(player_name)
		if pid is None:
			return {"name": player_name, "found": False}
		cur = self.conn.cursor()
		cur.execute("SELECT name, sex, weight FROM players WHERE id=?", (pid,))
		name, sex, weight = cur.fetchone()
		cur.execute("SELECT context_hand AS hand, mu, sigma, (mu - 3*sigma) AS conservative, updated_at FROM ratings WHERE player_id=? ORDER BY updated_at DESC", (pid,))
		ratings = [dict(row) for row in cur.fetchall()]
		cur.execute("SELECT e.id, e.event_type, e.name, e.date FROM events e JOIN supermatches s ON s.event_id=e.id WHERE s.a_player_id=? OR s.b_player_id=? ORDER BY e.id DESC", (pid, pid))
		sm = [dict(row) for row in cur.fetchall()]
		cur.execute("SELECT e.id, e.event_type, e.name, e.date, t.placement FROM events e JOIN tournament_results t ON t.event_id=e.id WHERE t.player_id=? ORDER BY e.id DESC", (pid,))
		tr = [dict(row) for row in cur.fetchall()]
		return {"name": name, "sex": sex, "weight": weight, "found": True, "ratings": ratings, "supermatches": sm, "tournaments": tr}


class TrueSkillService:
	def __init__(self, draw_probability: float = 0.01) -> None:
		# Prefer high-precision backend to avoid floating point issues on extreme cases
		self.env = ts.TrueSkill(draw_probability=draw_probability)
		try:
			ts.setup(
				mu=self.env.mu,
				sigma=self.env.sigma,
				beta=self.env.beta,
				tau=self.env.tau,
				draw_probability=self.env.draw_probability,
				backend='mpmath',
			)
		except Exception:
			ts.setup(
				mu=self.env.mu,
				sigma=self.env.sigma,
				beta=self.env.beta,
				tau=self.env.tau,
				draw_probability=self.env.draw_probability,
			)

	def rate_tournament(self, ordered_players: Sequence[ts.Rating], ranks: Sequence[int]) -> List[ts.Rating]:
		teams = [{i: r} for i, r in enumerate(ordered_players)]
		new_ratings_list = ts.rate(teams, ranks=ranks)  # type: ignore[arg-type]
		flattened: List[ts.Rating] = [None] * len(ordered_players)  # type: ignore[assignment]
		for d in new_ratings_list:
			idx, rating = next(iter(d.items()))
			flattened[idx] = rating
		return flattened

	def rate_game(self, a: ts.Rating, b: ts.Rating, a_won: bool) -> Tuple[ts.Rating, ts.Rating]:
		if a_won:
			na, nb = ts.rate_1vs1(a, b, min_delta=1e-9)
		else:
			nb, na = ts.rate_1vs1(b, a, min_delta=1e-9)
		return na, nb

	def conservative(self, r: ts.Rating) -> float:
		return float(r.mu - 3.0 * r.sigma)

	def rate_weighted_single(self, a: ts.Rating, b: ts.Rating, a_won: bool, weight: float, guarantee_flip: bool = False) -> Tuple[ts.Rating, ts.Rating]:
		# Base single pin update
		na1, nb1 = self.rate_game(a, b, a_won=a_won)
		dmu_a, dsig_a = (na1.mu - a.mu), (na1.sigma - a.sigma)
		dmu_b, dsig_b = (nb1.mu - b.mu), (nb1.sigma - b.sigma)
		scale = max(0.0, float(weight))
		if guarantee_flip:
			# Ensure winner's conservative score exceeds loser's
			C = (a.mu - 3.0 * a.sigma) - (b.mu - 3.0 * b.sigma)
			D = (dmu_a - 3.0 * dsig_a) - (dmu_b - 3.0 * dsig_b)
			if D > 0:
				required = (0.0001 - C) / D
				if required > scale:
					scale = required
		# Per-event sigma clamps: keep sigma from collapsing in one event
		MIN_SIGMA = 2.5
		MAX_DROP = 0.25  # at most 25% decrease per event
		cand_mu_a = a.mu + scale * dmu_a
		cand_mu_b = b.mu + scale * dmu_b
		cand_sig_a = a.sigma + scale * dsig_a
		cand_sig_b = b.sigma + scale * dsig_b
		floor_a = max(MIN_SIGMA, a.sigma * (1.0 - MAX_DROP))
		floor_b = max(MIN_SIGMA, b.sigma * (1.0 - MAX_DROP))
		new_a = ts.Rating(mu=cand_mu_a, sigma=max(floor_a, cand_sig_a))
		new_b = ts.Rating(mu=cand_mu_b, sigma=max(floor_b, cand_sig_b))
		return new_a, new_b


def normalize_hand(v: str) -> str:
	v = (v or "").strip().upper()
	if v in ("R", "RIGHT", "RH"): return "RH"
	if v in ("L", "LEFT", "LH"): return "LH"
	raise ValueError(f"Invalid hand: {v}")


def resolve_tournament(
		repo: RatingsRepository,
		logic: TrueSkillService,
		name: str,
		date_iso: str,
		hand: str,
		placements: Sequence[str],
	) -> int:
	"""
	Update ratings for a tournament given a finishing order (1st..N). Returns event_id.
	If exactly two names are provided, treat as a single pin (winner first, loser second).
	"""
	hand_n = normalize_hand(hand)
	event_id = repo.create_event("tournament", name, date_iso, hand_n)
	player_ids = [repo.get_or_create_player(n) for n in placements]
	if len(player_ids) == 2:
		# single pin: first beats second
		winner_id, loser_id = player_ids[0], player_ids[1]
		w = repo.get_rating(winner_id, hand_n, logic.env)
		l = repo.get_rating(loser_id, hand_n, logic.env)
		new_w, new_l = logic.rate_weighted_single(w, l, a_won=True, weight=1.0, guarantee_flip=False)
		repo.update_rating(winner_id, hand_n, new_w)
		repo.update_rating(loser_id, hand_n, new_l)
		repo.add_tournament_result(event_id, winner_id, 1)
		repo.add_tournament_result(event_id, loser_id, 2)
		return event_id
	ratings = [repo.get_rating(pid, hand_n, logic.env) for pid in player_ids]
	ranks = list(range(len(placements)))
	new_ratings = logic.rate_tournament(ratings, ranks)
	for pid, nr in zip(player_ids, new_ratings):
		repo.update_rating(pid, hand_n, nr)
	for place, pid in enumerate(player_ids, start=1):
		repo.add_tournament_result(event_id, pid, place)
	return event_id


def resolve_supermatch(
		repo: RatingsRepository,
		logic: TrueSkillService,
		name: str,
		date_iso: str,
		hand: str,
		a_name: str,
		b_name: str,
		a_wins: int,
		b_wins: int,
		best_of: Optional[int] = None,
	) -> int:
	"""
	Update ratings for a supermatch as a single weighted 1vs1 result.
	Winner receives a full heavy update (3x tournament pin). If the winner's
	ending conservative is below the loser's starting conservative, nudge only
	the winner's mu upward just enough to exceed the loser's starting value.
	If not a sweep, apply a tiny reverse win for the loser proportional to losses,
	clamped so the winner remains reasonable and numerical issues are avoided.
	Returns event_id.
	"""
	hand_n = normalize_hand(hand)
	event_id = repo.create_event("supermatch", name, date_iso, hand_n)
	a_id = repo.get_or_create_player(a_name)
	b_id = repo.get_or_create_player(b_name)
	a_rating = repo.get_rating(a_id, hand_n, logic.env)
	b_rating = repo.get_rating(b_id, hand_n, logic.env)
	# Defaults: 3-0 if nothing provided
	aw = int(a_wins or 0)
	bw = int(b_wins or 0)
	if (aw + bw) == 0:
		aw, bw = 3, 0
	# Decide winner; avoid perfect tie
	if aw == bw:
		aw += 1
	winner_is_a = aw > bw
	wins = max(aw, bw)
	losses = min(aw, bw)
	# Record loser's starting conservative score
	if winner_is_a:
		loser_start_cons = logic.conservative(b_rating)
	else:
		loser_start_cons = logic.conservative(a_rating)
	# 1) Big winner update: full weight 3.0 (no force-flip scaling)
	BASE_WEIGHT = 3.0
	if winner_is_a:
		w0, l0 = a_rating, b_rating
		new_w, new_l = logic.rate_weighted_single(w0, l0, a_won=True, weight=BASE_WEIGHT, guarantee_flip=False)
	else:
		w0, l0 = b_rating, a_rating
		new_w, new_l = logic.rate_weighted_single(w0, l0, a_won=True, weight=BASE_WEIGHT, guarantee_flip=False)
	# 2) Tiny reverse win for the loser if not a sweep (losses>0)
	if losses > 0 and wins > 0:
		try:
			EPS = 0.15
			tiny_weight = BASE_WEIGHT * EPS * (losses / float(wins))
			na1, nb1 = logic.rate_game(new_l, new_w, a_won=True)
			dmu_l, dsig_l = (na1.mu - new_l.mu), (na1.sigma - new_l.sigma)
			dmu_w, dsig_w = (nb1.mu - new_w.mu), (nb1.sigma - new_w.sigma)
			C0 = (new_w.mu - 3.0 * new_w.sigma) - (new_l.mu - 3.0 * new_l.sigma)
			D = (dmu_w - 3.0 * dsig_w) - (dmu_l - 3.0 * dsig_l)
			if D < 0:
				c_max = max(0.0, (C0 - 0.0001) / (-D))
				tiny_weight = min(tiny_weight, c_max)
			if tiny_weight > 0:
				na, nb = logic.rate_weighted_single(new_l, new_w, a_won=True, weight=tiny_weight, guarantee_flip=False)
				new_l, new_w = na, nb
		except Exception:
			pass
	# 3) Winner-only epsilon nudge to exceed loser's starting conservative if needed
	eps = 0.001
	new_w_cons = new_w.mu - 3.0 * new_w.sigma
	if new_w_cons <= loser_start_cons:
		delta = (loser_start_cons + eps) - new_w_cons
		new_w = ts.Rating(mu=new_w.mu + delta, sigma=new_w.sigma)
	# Persist results
	if winner_is_a:
		repo.update_rating(a_id, hand_n, new_w)
		repo.update_rating(b_id, hand_n, new_l)
	else:
		repo.update_rating(b_id, hand_n, new_w)
		repo.update_rating(a_id, hand_n, new_l)
	repo.add_supermatch_row(event_id, a_id, b_id, aw, bw, int(best_of or (aw + bw)))
	return event_id


def parse_event_json(path: str) -> Dict:
	with open(path, "r", encoding="utf-8") as f:
		return json.load(f)


def handle_ingest_json(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	logic = TrueSkillService()
	payload = parse_event_json(args.file)
	etype = str(payload.get("type", "")).strip().lower()
	name = str(payload.get("name", "Unnamed Event"))
	date_iso = str(payload.get("date", dt.date.today().isoformat()))
	hand = payload.get("hand", "RH")
	if etype == "tournament":
		placements_raw = payload.get("placements", [])
		if all(isinstance(x, str) for x in placements_raw):
			placements: List[str] = [str(x).strip() for x in placements_raw if str(x).strip()]
		else:
			objs = [x for x in placements_raw if isinstance(x, dict) and str(x.get("name", "")).strip()]
			objs.sort(key=lambda o: int(o.get("place", 999999)))
			placements = [str(o["name"]).strip() for o in objs]
		if not placements:
			raise ValueError("No placements found in tournament payload")
		resolve_tournament(repo, logic, name, date_iso, hand, placements)
		print(f"Ingested tournament '{name}' with {len(placements)} placements")
	elif etype == "supermatch":
		a_name = str(payload.get("a") or payload.get("player_a") or payload.get("blue") or "").strip()
		b_name = str(payload.get("b") or payload.get("player_b") or payload.get("red") or "").strip()
		if not a_name or not b_name:
			raise ValueError("Supermatch payload must include players 'a' and 'b'")
		score = payload.get("score")
		a_wins = payload.get("a_wins")
		b_wins = payload.get("b_wins")
		aw = int(a_wins) if a_wins is not None else 0
		bw = int(b_wins) if b_wins is not None else 0
		if isinstance(score, str) and "-" in score:
			parts = score.replace(" ", "").split("-")
			if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
				aw, bw = int(parts[0]), int(parts[1])
		best_of = payload.get("best_of")
		resolve_supermatch(repo, logic, name, date_iso, hand, a_name, b_name, aw, bw, best_of)
		print(f"Ingested supermatch '{a_name}' vs '{b_name}' ({aw}-{bw})")
	else:
		raise ValueError(f"Unknown event type: {etype}")


def handle_tournament_cli(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	logic = TrueSkillService()
	# Support single-pin form via --winner/--loser, else placements list (optionally two names)
	if args.winner and args.loser:
		placements = [args.winner.strip(), args.loser.strip()]
	else:
		placements = [n.strip() for n in (args.placements or []) if n.strip()]
	if not placements:
		raise ValueError("Provide --winner and --loser, or a placements list")
	resolve_tournament(
		repo,
		logic,
		name=args.name,
		date_iso=args.date or dt.date.today().isoformat(),
		hand=args.hand,
		placements=placements,
	)
	print(f"Tournament '{args.name}' processed with {len(placements)} placement(s)")


def handle_supermatch_cli(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	logic = TrueSkillService()
	aw = int(args.a_wins) if args.a_wins is not None else 0
	bw = int(args.b_wins) if args.b_wins is not None else 0
	best_of = args.best_of or (aw + bw if (aw + bw) > 0 else None)
	resolve_supermatch(
		repo,
		logic,
		name=args.name,
		date_iso=args.date or dt.date.today().isoformat(),
		hand=args.hand,
		a_name=args.player_a,
		b_name=args.player_b,
		a_wins=aw,
		b_wins=bw,
		best_of=best_of,
	)
	print(f"Supermatch '{args.player_a}' vs '{args.player_b}' processed ({aw}-{bw})")


def handle_show_ratings(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	ratings = repo.fetch_ratings(args.hand)
	if not ratings:
		print("No ratings found.")
	else:
		print("Player Name | Hand | Mu | Sigma | Conservative | Updated At")
		print("-" * 90)
		for r in ratings:
			print(f"{r['name']:<15} | {r['hand']:<4} | {r['mu']:.2f} | {r['sigma']:.2f} | {r['conservative']:.2f} | {r['updated_at']}")


def handle_list_events(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	events = repo.fetch_events(args.limit)
	if not events:
		print("No events found.")
	else:
		print("Event ID | Type | Name | Date | Hand")
		print("-" * 80)
		for e in events:
			print(f"{e['id']:<8} | {e['event_type']:<10} | {e['name']:<20} | {e['date']} | {e['context_hand']:<4}")


def handle_show_player(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	overview = repo.fetch_player_overview(args.name)
	if not overview["found"]:
		print(f"Player '{args.name}' not found.")
	else:
		print(f"Player: {overview['name']}  Sex: {overview.get('sex') or ''}  Weight: {overview.get('weight') or ''}")
		print("-" * 50)
		print("Ratings:")
		if not overview["ratings"]:
			print("No ratings found.")
		else:
			print("Hand | Mu | Sigma | Conservative | Updated At")
			print("-" * 50)
			for r in overview["ratings"]:
				print(f"{r['hand']:<4} | {r['mu']:.2f} | {r['sigma']:.2f} | {r['conservative']:.2f} | {r['updated_at']}")
		print("\nEvent History:")
		if not overview["supermatches"] and not overview["tournaments"]:
			print("No events found.")
		else:
			print("Type | Event ID | Name | Date")
			print("-" * 50)
			for sm in overview["supermatches"]:
				print(f"Supermatch | {sm['id']:<8} | {sm['name']:<20} | {sm['date']}")
			for tr in overview["tournaments"]:
				print(f"Tournament | {tr['id']:<8} | {tr['name']:<20} | {tr['date']}")


def handle_import_xlsx(conn: sqlite3.Connection, args: argparse.Namespace) -> None:
	repo = RatingsRepository(conn)
	logic = TrueSkillService()
	from openpyxl import load_workbook
	path = args.file
	wb = load_workbook(path, data_only=True)

	def find_sheet(candidates: List[str]) -> Optional[str]:
		lower_map = {sh.title.lower(): sh.title for sh in wb.worksheets}
		for cand in candidates:
			for key, title in lower_map.items():
				if cand in key:
					return title
		return None

	def parse_weights(title: Optional[str], sex: Optional[str]) -> Dict[str, float]:
		if not title or title not in wb.sheetnames:
			return {}
		ws = wb[title]
		# Expect header row with 'name' and 'weight'
		head = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))[0:3]]
		name_idx = None
		w_idx = None
		for i, h in enumerate(head):
			if h in ('name', 'athlete', 'puller'):
				name_idx = i
			if 'weight' in h:
				w_idx = i
		if name_idx is None or w_idx is None:
			# fallback: assume first two cols
			name_idx, w_idx = 0, 1
		out: Dict[str, float] = {}
		for row in ws.iter_rows(min_row=2):
			name = str((row[name_idx].value or '')).strip()
			if not name:
				continue
			w_raw = row[w_idx].value
			try:
				w = float(str(w_raw).strip()) if w_raw not in (None, '') else None
			except Exception:
				w = None
			if w is not None:
				out[name] = w
			# also set sex if provided
			pid = repo.get_or_create_player(name)
			if sex is not None:
				repo.set_player_attrs(pid, sex=sex)
			if w is not None:
				repo.set_player_attrs(pid, weight=w)
		return out

	def parse_rankings(title: Optional[str]) -> Dict[str, List[str]]:
		if not title or title not in wb.sheetnames:
			return {'RH': [], 'LH': []}
		ws = wb[title]
		# Find header cells for left/right
		head = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))[0:6]]
		lh_idx = None
		rh_idx = None
		for idx, h in enumerate(head):
			if 'left' in h or h == 'lh':
				lh_idx = idx
			if 'right' in h or h == 'rh':
				rh_idx = idx
		# fallback assume first two columns are LH, RH
		if lh_idx is None:
			lh_idx = 0
		if rh_idx is None:
			rh_idx = 1 if lh_idx == 0 else 0
		lists = {'LH': [], 'RH': []}
		for row in ws.iter_rows(min_row=2):
			def cell_str(idx):
				try:
					return str((row[idx].value or '')).strip()
				except Exception:
					return ''
			ln = cell_str(lh_idx)
			rn = cell_str(rh_idx)
			if ln:
				lists['LH'].append(ln)
			if rn:
				lists['RH'].append(rn)
		# dedupe preserving order
		for k in ('LH', 'RH'):
			seen = set(); uniq = []
			for n in lists[k]:
				if n not in seen:
					seen.add(n); uniq.append(n)
			lists[k] = uniq
		return lists

	# Guess sheets if not provided
	men_ranks = args.men_rankings or find_sheet(['rankings_fals', 'rankings men', 'rankings'])
	women_ranks = args.women_rankings or find_sheet(['rankings_f', 'rankings women', 'rankings_female'])
	men_weights = args.men_weights or find_sheet(['weights men', 'weights'])
	women_weights = args.women_weights or find_sheet(['weights_f', 'weights women'])

	# Import weights first (also sets sex)
	parse_weights(men_weights, sex='M')
	parse_weights(women_weights, sex='F')

	mode = (args.mode or 'adhoc').strip().lower()
	if mode == 'tournament':
		# Import rankings as virtual tournaments
		for title, sex in ((men_ranks, 'M'), (women_ranks, 'F')):
			if not title:
				continue
			lists = parse_rankings(title)
			for hand_key in ('RH', 'LH'):
				placements = lists[hand_key]
				if not placements:
					continue
				# Ensure players exist and set sex if available
				for n in placements:
					pid = repo.get_or_create_player(n)
					repo.set_player_attrs(pid, sex=sex)
				resolve_tournament(repo, logic, name=f"Seed from {title} {hand_key}", date_iso=dt.date.today().isoformat(), hand=hand_key, placements=placements)
	else:
		# Ad-hoc seeding: assign mu linearly and sigma by quartile per weight class, preserving conservative order
		def weight_class(lbs: Optional[float], classes: List[str]) -> str:
			# Default unknown or -1 to top (super heavy) class
			if lbs is None:
				return classes[-1]
			try:
				w = float(lbs)
			except Exception:
				return classes[-1]
			if w == -1:
				return classes[-1]
			thresholds = []
			for c in classes:
				if c.endswith('+'):
					continue
				try:
					thresholds.append(float(c))
				except Exception:
					pass
			for t in thresholds:
				if w <= t:
					return str(int(t))
			return classes[-1]

		men_classes = ['154','176','198','220','242','243+']
		women_classes = ['143','144+']

		# Build weight maps from DB
		cur = conn.cursor()
		cur.execute("SELECT name, sex, weight FROM players")
		rows = list(cur.fetchall())
		name_to_weight = {str(r['name']).strip(): (('M' if (r['sex'] or '').upper().startswith('M') else ('F' if (r['sex'] or '').upper().startswith('F') else None)), (float(r['weight']) if r['weight'] is not None else None)) for r in rows}

		def class_quartiles(placements: List[str], sex: str) -> Dict[str, str]:
			# returns athlete->sigma_tier label based on quartile within their weight class
			# tiers: 'very','confident','semi','not' with approximately equal counts per class
			classes = men_classes if sex == 'M' else women_classes
			by_cls: Dict[str, List[str]] = {}
			for n in placements:
				_, w = name_to_weight.get(n, (sex, None))
				cls = weight_class(w, classes)
				by_cls.setdefault(cls, []).append(n)
			out: Dict[str, str] = {}
			for cls, lst in by_cls.items():
				n = len(lst)
				if n == 0:
					continue
				# chunk sizes using ceiling to allocate close to 25% per bucket
				q = math.ceil(n / 4)
				b1 = min(q, n)
				b2 = min(q, max(0, n - b1))
				b3 = min(q, max(0, n - b1 - b2))
				b4 = max(0, n - b1 - b2 - b3)
				idx = 0
				for k in range(b1): out[lst[idx+k]] = 'very'
				idx += b1
				for k in range(b2): out[lst[idx+k]] = 'confident'
				idx += b2
				for k in range(b3): out[lst[idx+k]] = 'semi'
				idx += b3
				for k in range(b4): out[lst[idx+k]] = 'not'
			return out

		SIGMA_BY_TIER = {'very': 3.0, 'confident': 4.0, 'semi': 5.5, 'not': 6.5}
		MU_STEP = 1.0
		CONS_MARGIN_WOMEN = 0.5  # proportionally smaller margin with tighter mu spacing

		def seed_hand(men_list: List[str], women_list: List[str], hand_key: str) -> None:
			N_m = len(men_list)
			N_w = len(women_list)
			N_total = N_m + N_w
			if N_total == 0:
				return
			# Top mu so that the last overall mu = 25
			mu_top = 25.0 + MU_STEP * (N_total - 1)
			# MEN: sigma tiers by quartile within class, then smooth to be non-decreasing along list
			tiers_m = class_quartiles(men_list, 'M')
			sigma_m: List[float] = []
			for i, name in enumerate(men_list):
				pid = repo.get_or_create_player(name)
				repo.set_player_attrs(pid, sex='M')
				sigma_m.append(SIGMA_BY_TIER[tiers_m.get(name, 'not')])
			for i in range(1, len(sigma_m)):
				if sigma_m[i] < sigma_m[i-1]:
					sigma_m[i] = sigma_m[i-1]
			# Apply ratings for men
			for i, name in enumerate(men_list):
				mu = mu_top - MU_STEP * i
				repo.update_rating(repo.get_or_create_player(name), hand_key, ts.Rating(mu=mu, sigma=sigma_m[i]))
			# Track bottom man's conservative
			if N_m > 0:
				mu_bottom_man = mu_top - MU_STEP * (N_m - 1)
				sigma_bottom_man = sigma_m[-1]
				cons_bottom_man = mu_bottom_man - 3.0 * sigma_bottom_man
			else:
				cons_bottom_man = float('-inf')
			# WOMEN: top woman confident (4.0), others not (6.5)
			if N_w == 0:
				return
			# Default mu progression for women continues the sequence
			mu_w0_default = mu_top - MU_STEP * N_m
			sigma_top_w = SIGMA_BY_TIER['confident']
			# Place top woman just below bottom man by CONS_MARGIN_WOMEN
			if N_m > 0:
				cons_target = cons_bottom_man - CONS_MARGIN_WOMEN
				mu_w0 = cons_target + 3.0 * sigma_top_w
			else:
				mu_w0 = mu_w0_default
			# Assign women ratings
			for j, name in enumerate(women_list):
				pid = repo.get_or_create_player(name)
				repo.set_player_attrs(pid, sex='F')
				if j == 0:
					sigma = sigma_top_w
					mu = mu_w0
				else:
					sigma = SIGMA_BY_TIER['not']
					mu = (mu_w0 - MU_STEP * j)
				repo.update_rating(pid, hand_key, ts.Rating(mu=mu, sigma=sigma))

		# Build lists and seed per hand
		men_lists = parse_rankings(men_ranks) if men_ranks else {'RH': [], 'LH': []}
		women_lists = parse_rankings(women_ranks) if women_ranks else {'RH': [], 'LH': []}
		for hand_key in ('RH', 'LH'):
			seed_hand(men_lists[hand_key], women_lists[hand_key], hand_key)

	print("Import complete.")


def make_parser() -> argparse.ArgumentParser:
	p = argparse.ArgumentParser(description="Resolve tournament and supermatch outcomes with TrueSkill")
	p.add_argument("--db", default=DB_DEFAULT_PATH, help="Path to SQLite DB (default: backend/data/ratings.db)")
	sub = p.add_subparsers(dest="cmd", required=True)

	p_ingest = sub.add_parser("ingest", help="Ingest event from JSON payload")
	p_ingest.add_argument("--file", required=True, help="Path to JSON file")
	p_ingest.set_defaults(func=handle_ingest_json)

	p_t = sub.add_parser("tournament", help="Record a tournament by placements or single pin")
	p_t.add_argument("--name", required=True)
	p_t.add_argument("--date", default=None, help="ISO date (YYYY-MM-DD), default today")
	p_t.add_argument("--hand", required=True, choices=["RH", "LH", "right", "left", "r", "l"])
	p_t.add_argument("--winner", default=None, help="Winner name (single pin mode)")
	p_t.add_argument("--loser", default=None, help="Loser name (single pin mode)")
	p_t.add_argument("placements", nargs="*", help="Ordered names: first is 1st (optional if using --winner/--loser)")
	p_t.set_defaults(func=handle_tournament_cli)

	p_s = sub.add_parser("supermatch", help="Record a supermatch best-of result")
	p_s.add_argument("--name", required=True)
	p_s.add_argument("--date", default=None)
	p_s.add_argument("--hand", required=True, choices=["RH", "LH", "right", "left", "r", "l"])
	p_s.add_argument("--player_a", required=True)
	p_s.add_argument("--player_b", required=True)
	p_s.add_argument("--a_wins", type=int, required=False, help="Optional A wins; defaults to 3 when omitted with B=0")
	p_s.add_argument("--b_wins", type=int, required=False, help="Optional B wins; defaults to 0 when omitted with A=3")
	p_s.add_argument("--best_of", type=int, default=None)
	p_s.set_defaults(func=handle_supermatch_cli)

	p_r = sub.add_parser("ratings", help="Show ratings (optionally filter by hand)")
	p_r.add_argument("--hand", default=None, choices=[None, "RH", "LH", "right", "left", "r", "l"], help="Filter by hand")
	p_r.set_defaults(func=handle_show_ratings)

	p_e = sub.add_parser("events", help="List recent events")
	p_e.add_argument("--limit", type=int, default=20)
	p_e.set_defaults(func=handle_list_events)

	p_p = sub.add_parser("player", help="Show a player's ratings and event history")
	p_p.add_argument("--name", required=True)
	p_p.set_defaults(func=handle_show_player)

	p_imp = sub.add_parser("import-xlsx", help="Initialize players and ratings from an Excel rankings export")
	p_imp.add_argument("--file", required=True, help="Path to .xlsx file")
	p_imp.add_argument("--men-rankings", default=None, help="Sheet name for men's rankings")
	p_imp.add_argument("--women-rankings", default=None, help="Sheet name for women's rankings")
	p_imp.add_argument("--men-weights", default=None, help="Sheet name for men's weights")
	p_imp.add_argument("--women-weights", default=None, help="Sheet name for women's weights")
	p_imp.add_argument("--mode", default="adhoc", choices=["adhoc","tournament"], help="Seeding mode: adhoc or tournament-based")
	p_imp.set_defaults(func=handle_import_xlsx)

	return p


def main() -> None:
	args = make_parser().parse_args()
	conn = open_db(args.db)
	init_db(conn)
	args.func(conn, args)


if __name__ == "__main__":
	main() 